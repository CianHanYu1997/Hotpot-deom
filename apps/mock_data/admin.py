from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.db.models import Sum, Count
from django.utils import timezone
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from .models import Store, Revenue, Ingredient, Order, OrderItem, RevenueLog


@admin.register(Store)
class StoreAdmin(admin.ModelAdmin):
    list_display = ['name', 'location', 'get_region_display', 'manager_name', 'total_revenue', 'orders_count']
    list_filter = ['region', 'created_at']
    search_fields = ['name', 'location', 'manager_name']
    readonly_fields = ['created_at']
    
    def total_revenue(self, obj):
        total = obj.revenues.filter(status='submitted').aggregate(Sum('amount'))['amount__sum']
        if total:
            return f"${total:,.2f}"
        return "$0.00"
    total_revenue.short_description = '總營收'
    total_revenue.admin_order_field = 'total_revenue'
    
    def orders_count(self, obj):
        count = obj.orders.count()
        return f"{count} 筆"
    orders_count.short_description = '訂單數'


@admin.register(Revenue)
class RevenueAdmin(admin.ModelAdmin):
    list_display = ['store_link', 'date', 'amount', 'amount_display', 'status_badge', 'status', 'created_by', 'updated_at']
    list_filter = ['status', 'date', 'store__region', 'store']
    search_fields = ['store__name', 'notes']
    date_hierarchy = 'date'
    readonly_fields = ['submitted_at', 'updated_at']
    list_editable = ['amount', 'status']
    
    fieldsets = (
        ('基本資訊', {
            'fields': ('store', 'date', 'amount', 'status')
        }),
        ('備註與記錄', {
            'fields': ('notes', 'created_by', 'updated_by')
        }),
        ('時間戳記', {
            'fields': ('submitted_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def store_link(self, obj):
        url = reverse('admin:mock_data_store_change', args=[obj.store.id])
        return format_html('<a href="{}">{}</a>', url, obj.store.name)
    store_link.short_description = '分店'
    
    def amount_display(self, obj):
        if obj.amount is not None:
            return format_html('<strong>${}</strong>', f"{obj.amount:,.2f}")
        return format_html('<strong>$0.00</strong>')
    amount_display.short_description = '金額'
    amount_display.admin_order_field = 'amount'
    
    def status_badge(self, obj):
        colors = {
            'pending': '#ffc107',
            'submitted': '#28a745'
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; border-radius: 3px; font-size: 11px;">{}</span>',
            colors.get(obj.status, '#6c757d'),
            obj.get_status_display()
        )
    status_badge.short_description = '狀態'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('store')


@admin.register(Ingredient)
class IngredientAdmin(admin.ModelAdmin):
    list_display = ['name', 'unit', 'category', 'orders_count']
    list_filter = ['category', 'unit']
    search_fields = ['name']
    
    def orders_count(self, obj):
        count = OrderItem.objects.filter(ingredient=obj).count()
        return f"{count} 次"
    orders_count.short_description = '被訂購次數'


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ['ingredient_info', 'subtotal_display']
    fields = ['ingredient', 'quantity', 'ingredient_info', 'subtotal_display']
    
    def ingredient_info(self, obj):
        if obj.ingredient:
            return format_html(
                '<span style="color: #666; font-size: 11px;">{} | {}</span>',
                obj.ingredient.category,
                obj.ingredient.unit
            )
        return ""
    ingredient_info.short_description = '食材資訊'
    
    def subtotal_display(self, obj):
        if obj.ingredient and obj.quantity:
            return format_html(
                '<strong>{} × {} {}</strong>',
                obj.ingredient.name,
                obj.quantity,
                obj.ingredient.unit
            )
        return ""
    subtotal_display.short_description = '訂購明細'


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ['store_link', 'date', 'items_count', 'items_preview', 'completion_status', 'submitted_at']
    list_filter = ['is_completed', 'date', 'store__region', 'store']
    search_fields = ['store__name']
    date_hierarchy = 'date'
    inlines = [OrderItemInline]
    readonly_fields = ['submitted_at']
    actions = ['export_orders_to_excel']
    
    def store_link(self, obj):
        url = reverse('admin:mock_data_store_change', args=[obj.store.id])
        return format_html('<a href="{}">{}</a>', url, obj.store.name)
    store_link.short_description = '分店'
    
    def items_count(self, obj):
        count = obj.items.count()
        return f"{count} 項目"
    items_count.short_description = '項目數'
    
    def items_preview(self, obj):
        items = obj.items.all()[:3]
        if not items:
            return "無項目"
        
        preview = ", ".join([f"{item.ingredient.name}({item.quantity})" for item in items])
        total_count = obj.items.count()
        
        if total_count > 3:
            preview += f" ...等{total_count}項"
        
        return format_html('<span title="{}">{}</span>', preview, preview[:50])
    items_preview.short_description = '叫菜內容'
    
    def completion_status(self, obj):
        if obj.is_completed:
            return format_html(
                '<span style="color: green;">✓ 已完成</span>'
            )
        else:
            return format_html(
                '<span style="color: orange;">⏳ 處理中</span>'
            )
    completion_status.short_description = '狀態'
    
    def export_orders_to_excel(self, request, queryset):
        """匯出選定的訂單到Excel"""
        workbook = Workbook()
        
        # 訂單摘要工作表
        summary_sheet = workbook.active
        summary_sheet.title = "訂單摘要"
        
        # 設定標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 摘要表標題
        summary_headers = ['分店', '地區', '日期', '項目數', '狀態', '提交時間']
        for col, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # 填入摘要資料
        for row, order in enumerate(queryset.select_related('store'), 2):
            summary_sheet.cell(row=row, column=1, value=order.store.name).border = border
            summary_sheet.cell(row=row, column=2, value=order.store.get_region_display()).border = border
            summary_sheet.cell(row=row, column=3, value=order.date.strftime('%Y-%m-%d')).border = border
            summary_sheet.cell(row=row, column=4, value=order.items.count()).border = border
            summary_sheet.cell(row=row, column=5, value="已完成" if order.is_completed else "處理中").border = border
            summary_sheet.cell(row=row, column=6, value=order.submitted_at.strftime('%Y-%m-%d %H:%M')).border = border
        
        # 調整欄位寬度
        for col in range(1, 7):
            summary_sheet.column_dimensions[summary_sheet.cell(row=1, column=col).column_letter].width = 15
        
        # 詳細內容工作表
        detail_sheet = workbook.create_sheet("訂單明細")
        
        detail_headers = ['分店', '地區', '日期', '食材名稱', '分類', '數量', '單位', '狀態']
        for col, header in enumerate(detail_headers, 1):
            cell = detail_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # 填入詳細資料
        detail_row = 2
        for order in queryset.select_related('store').prefetch_related('items__ingredient'):
            for item in order.items.all():
                detail_sheet.cell(row=detail_row, column=1, value=order.store.name).border = border
                detail_sheet.cell(row=detail_row, column=2, value=order.store.get_region_display()).border = border
                detail_sheet.cell(row=detail_row, column=3, value=order.date.strftime('%Y-%m-%d')).border = border
                detail_sheet.cell(row=detail_row, column=4, value=item.ingredient.name).border = border
                detail_sheet.cell(row=detail_row, column=5, value=item.ingredient.category).border = border
                detail_sheet.cell(row=detail_row, column=6, value=item.quantity).border = border
                detail_sheet.cell(row=detail_row, column=7, value=item.ingredient.unit).border = border
                detail_sheet.cell(row=detail_row, column=8, value="已完成" if order.is_completed else "處理中").border = border
                detail_row += 1
        
        # 調整欄位寬度
        for col in range(1, 9):
            detail_sheet.column_dimensions[detail_sheet.cell(row=1, column=col).column_letter].width = 15
        
        # 準備回應
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="叫菜訂單_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        workbook.save(response)
        return response
        
    export_orders_to_excel.short_description = "匯出選定訂單到Excel"


@admin.register(RevenueLog)
class RevenueLogAdmin(admin.ModelAdmin):
    list_display = ['store', 'date', 'action_badge', 'amount_change', 'status_change', 'changed_by', 'changed_at']
    list_filter = ['action', 'date', 'store__region', 'store']
    search_fields = ['store__name', 'notes', 'changed_by']
    date_hierarchy = 'changed_at'
    readonly_fields = ['changed_at']
    
    def action_badge(self, obj):
        colors = {
            'create': '#007bff',
            'update': '#ffc107',
            'delete': '#dc3545',
            'submit': '#28a745'
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px;">{}</span>',
            colors.get(obj.action, '#6c757d'),
            obj.get_action_display()
        )
    action_badge.short_description = '操作'
    
    def amount_change(self, obj):
        if obj.old_amount and obj.new_amount:
            return f"${obj.old_amount:,.2f} → ${obj.new_amount:,.2f}"
        elif obj.new_amount:
            return f"新增 ${obj.new_amount:,.2f}"
        elif obj.old_amount:
            return f"刪除 ${obj.old_amount:,.2f}"
        return "-"
    amount_change.short_description = '金額變更'
    
    def status_change(self, obj):
        if obj.old_status and obj.new_status:
            return f"{obj.old_status} → {obj.new_status}"
        elif obj.new_status:
            return obj.new_status
        return "-"
    status_change.short_description = '狀態變更'


# 自訂管理介面標題
admin.site.site_header = "餐廳管理系統"
admin.site.site_title = "餐廳管理"
admin.site.index_title = "後台管理"
