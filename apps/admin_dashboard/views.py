from django.shortcuts import render, get_object_or_404, redirect
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Sum, Avg
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.urls import reverse
import json
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import timedelta, datetime
from apps.mock_data.models import Store, Revenue, Order, OrderItem, Ingredient, RevenueLog

# 自定義 JSON 編碼器，處理 Decimal 類型


class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super(DecimalEncoder, self).default(obj)


class DashboardView(View):
    """Main dashboard view for administrators"""

    def get(self, request):
        today = timezone.now().date()

        # Get all stores
        stores = Store.objects.all()

        # Get today's total revenue
        today_revenue = Revenue.objects.filter(
            date=today).aggregate(total=Sum('amount'))

        # Get yesterday's total revenue for comparison
        yesterday = today - timedelta(days=1)
        yesterday_revenue = Revenue.objects.filter(
            date=yesterday).aggregate(total=Sum('amount'))

        # Calculate revenue change percentage
        revenue_change = 0
        if yesterday_revenue['total'] and today_revenue['total']:
            revenue_change = (
                (today_revenue['total'] - yesterday_revenue['total']) / yesterday_revenue['total']) * 100

        # Get stores that haven't completed orders today
        incomplete_orders = []
        for store in stores:
            today_order = Order.objects.filter(store=store, date=today).first()
            if not today_order or not today_order.is_completed:
                incomplete_orders.append(store)

        # Get last 7 days revenue data for chart
        last_7_days = []
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            daily_revenue = Revenue.objects.filter(
                date=date).aggregate(total=Sum('amount'))
            last_7_days.append({
                'date': date.strftime('%m/%d'),
                # 將 Decimal 轉換為 float
                'revenue': float(daily_revenue['total'] or 0)
            })

        context = {
            'stores': stores,
            'today_revenue': today_revenue['total'] or 0,
            'yesterday_revenue': yesterday_revenue['total'] or 0,
            'revenue_change': revenue_change,
            'incomplete_orders': incomplete_orders,
            # 使用自定義編碼器
            'last_7_days': json.dumps(last_7_days, cls=DecimalEncoder),
            'today': today,
        }
        return render(request, 'admin_dashboard/dashboard.html', context)


class RevenueManageView(View):
    """Enhanced revenue statistics view with store filtering and regional analysis"""

    def get(self, request):
        # Get date range and filter parameters
        today = timezone.now().date()
        start_date_str = request.GET.get(
            'start_date', (today - timedelta(days=30)).isoformat())
        end_date_str = request.GET.get('end_date', today.isoformat())
        store_id = request.GET.get('store_id')
        region = request.GET.get('region')
        status = request.GET.get('status')

        try:
            start_date = datetime.fromisoformat(start_date_str).date()
            end_date = datetime.fromisoformat(end_date_str).date()
        except ValueError:
            start_date = today - timedelta(days=30)
            end_date = today

        # Get filtered stores
        stores = Store.objects.all()
        if region:
            stores = stores.filter(region=region)
        if store_id:
            stores = stores.filter(id=store_id)

        # Get revenue data for filtered stores
        revenues_query = Revenue.objects.filter(
            date__range=[start_date, end_date]
        )
        if store_id:
            revenues_query = revenues_query.filter(store_id=store_id)
        elif region:
            revenues_query = revenues_query.filter(store__region=region)
        if status:
            revenues_query = revenues_query.filter(status=status)

        # Store-level revenue data
        store_revenues = []
        for store in stores:
            store_revenues_query = revenues_query.filter(store=store)
            revenues = store_revenues_query.order_by('date')

            total_revenue = revenues.aggregate(total=Sum('amount'))['total'] or 0
            avg_revenue = revenues.aggregate(avg=Avg('amount'))['avg'] or 0
            count = revenues.count()
            submitted_count = revenues.filter(status='submitted').count()
            pending_count = revenues.filter(status='pending').count()

            store_revenues.append({
                'store': store,
                'revenues': revenues,
                'total': total_revenue,
                'average': avg_revenue,
                'count': count,
                'submitted_count': submitted_count,
                'pending_count': pending_count,
            })

        # Regional summary data
        regional_data = []
        for region_code, region_name in Store.REGION_CHOICES:
            region_stores = Store.objects.filter(region=region_code)
            region_revenues = Revenue.objects.filter(
                store__region=region_code,
                date__range=[start_date, end_date]
            )
            if status:
                region_revenues = region_revenues.filter(status=status)
                
            total = region_revenues.aggregate(total=Sum('amount'))['total'] or 0
            count = region_revenues.count()
            avg = total / count if count > 0 else 0
            
            regional_data.append({
                'region_code': region_code,
                'region_name': region_name,
                'total': total,
                'average': avg,
                'count': count,
                'store_count': region_stores.count()
            })

        # Daily trend data for chart (last 7 days)
        daily_data = []
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            daily_revenues = Revenue.objects.filter(date=date)
            if store_id:
                daily_revenues = daily_revenues.filter(store_id=store_id)
            elif region:
                daily_revenues = daily_revenues.filter(store__region=region)
            if status:
                daily_revenues = daily_revenues.filter(status=status)
                
            daily_total = daily_revenues.aggregate(total=Sum('amount'))['total'] or 0
            daily_data.append({
                'date': date.strftime('%m/%d'),
                'revenue': float(daily_total)
            })

        # Calculate overall totals
        overall_total = sum(sr['total'] for sr in store_revenues)
        overall_avg = overall_total / len(store_revenues) if store_revenues else 0
        total_records = sum(sr['count'] for sr in store_revenues)

        context = {
            'store_revenues': store_revenues,
            'regional_data': regional_data,
            'daily_data': json.dumps(daily_data, cls=DecimalEncoder),
            'start_date': start_date,
            'end_date': end_date,
            'overall_total': overall_total,
            'overall_avg': overall_avg,
            'total_records': total_records,
            'stores': Store.objects.all(),
            'regions': Store.REGION_CHOICES,
            'status_choices': Revenue.STATUS_CHOICES,
            'selected_store': store_id,
            'selected_region': region,
            'selected_status': status,
        }
        return render(request, 'admin_dashboard/revenue.html', context)


class OrderManageView(View):
    """Order management view for administrators"""

    def get(self, request):
        today = timezone.now().date()

        # Get all stores
        stores = Store.objects.all()

        # Get today's orders for each store
        store_orders = []
        for store in stores:
            order = Order.objects.filter(store=store, date=today).first()
            store_orders.append({
                'store': store,
                'order': order,
                'items': order.items.all() if order else []
            })

        # Get all ingredients and their total quantities ordered today
        ingredients = Ingredient.objects.all()
        ingredient_totals = {}

        for ingredient in ingredients:
            total_quantity = OrderItem.objects.filter(
                order__date=today,
                ingredient=ingredient
            ).aggregate(total=Sum('quantity'))['total'] or 0

            ingredient_totals[ingredient.id] = {
                'ingredient': ingredient,
                'total_quantity': total_quantity
            }

        context = {
            'store_orders': store_orders,
            'ingredient_totals': ingredient_totals,
            'today': today,
        }
        return render(request, 'admin_dashboard/orders.html', context)


class ReportsView(View):
    """Reports generation view for administrators"""

    def get(self, request):
        today = timezone.now().date()

        context = {
            'today': today,
        }
        return render(request, 'admin_dashboard/reports.html', context)


def export_revenue_report(request):
    """Export revenue data to Excel"""
    # Get date range from query parameters
    today = timezone.now().date()
    start_date_str = request.GET.get(
        'start_date', (today - timedelta(days=30)).isoformat())
    end_date_str = request.GET.get('end_date', today.isoformat())

    try:
        start_date = datetime.fromisoformat(start_date_str).date()
        end_date = datetime.fromisoformat(end_date_str).date()
    except ValueError:
        start_date = today - timedelta(days=30)
        end_date = today

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "營收報表"

    # Add header row
    headers = ["日期", "分店", "營收金額"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD",
                                end_color="DDDDDD", fill_type="solid")

    # Add data rows
    revenues = Revenue.objects.filter(
        date__range=[start_date, end_date]).order_by('date', 'store__name')

    for row_num, revenue in enumerate(revenues, 2):
        ws.cell(row=row_num, column=1).value = revenue.date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=2).value = revenue.store.name
        ws.cell(row=row_num, column=3).value = float(revenue.amount)

    # Add summary row
    total_row = len(revenues) + 2
    ws.cell(row=total_row, column=1).value = "總計"
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=3).value = f"=SUM(C2:C{total_row-1})"
    ws.cell(row=total_row, column=3).font = Font(bold=True)

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    # Create the response with the Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=營收報表_{start_date}_to_{end_date}.xlsx'
    wb.save(response)

    return response


def export_orders_report(request):
    """Export ingredient orders data to Excel"""
    # Get date from query parameters
    date_str = request.GET.get('date', timezone.now().date().isoformat())

    try:
        report_date = datetime.fromisoformat(date_str).date()
    except ValueError:
        report_date = timezone.now().date()

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "叫菜報表"

    # Add header row
    headers = ["食材", "單位", "分類", "總數量"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD",
                                end_color="DDDDDD", fill_type="solid")

    # Get all ingredients and their total quantities ordered for the date
    ingredients = Ingredient.objects.all()
    row_num = 2

    for ingredient in ingredients:
        total_quantity = OrderItem.objects.filter(
            order__date=report_date,
            ingredient=ingredient
        ).aggregate(total=Sum('quantity'))['total'] or 0

        # Only include ingredients that were ordered
        if total_quantity > 0:
            ws.cell(row=row_num, column=1).value = ingredient.name
            ws.cell(row=row_num, column=2).value = ingredient.unit
            ws.cell(row=row_num, column=3).value = ingredient.category
            ws.cell(row=row_num, column=4).value = total_quantity
            row_num += 1

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10

    # Create the response with the Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=叫菜報表_{report_date}.xlsx'
    wb.save(response)

    return response


class RevenueCRUDView(View):
    """營收 CRUD 管理視圖"""
    
    def get(self, request):
        """顯示營收列表"""
        today = timezone.now().date()
        start_date = today - timedelta(days=30)
        
        # 獲取篩選參數
        store_id = request.GET.get('store_id')
        status = request.GET.get('status')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        # 建立查詢
        revenues = Revenue.objects.select_related('store').all()
        
        if store_id:
            revenues = revenues.filter(store_id=store_id)
        if status:
            revenues = revenues.filter(status=status)
        if start_date_str:
            try:
                start_date = datetime.fromisoformat(start_date_str).date()
                revenues = revenues.filter(date__gte=start_date)
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.fromisoformat(end_date_str).date()
                revenues = revenues.filter(date__lte=end_date)
            except ValueError:
                pass
        
        revenues = revenues.order_by('-date', '-updated_at')[:100]  # 限制顯示數量
        
        context = {
            'revenues': revenues,
            'stores': Store.objects.all(),
            'status_choices': Revenue.STATUS_CHOICES,
            'selected_store': store_id,
            'selected_status': status,
            'start_date': start_date_str,
            'end_date': end_date_str,
        }
        return render(request, 'admin_dashboard/revenue_crud.html', context)


class RevenueDetailView(View):
    """營收詳細資料和編輯視圖"""
    
    def get(self, request, revenue_id):
        """顯示營收詳細資料"""
        revenue = get_object_or_404(Revenue, id=revenue_id)
        logs = revenue.logs.all()[:20]  # 最近20筆記錄
        
        context = {
            'revenue': revenue,
            'logs': logs,
            'stores': Store.objects.all(),
            'status_choices': Revenue.STATUS_CHOICES,
        }
        return render(request, 'admin_dashboard/revenue_detail.html', context)
    
    def post(self, request, revenue_id):
        """更新營收資料"""
        revenue = get_object_or_404(Revenue, id=revenue_id)
        
        # 記錄變更前的值
        old_amount = revenue.amount
        old_status = revenue.status
        old_notes = revenue.notes
        
        # 更新資料
        new_amount = request.POST.get('amount')
        new_status = request.POST.get('status')
        new_notes = request.POST.get('notes', '')
        changed_by = request.POST.get('changed_by', '管理員')
        
        try:
            revenue.amount = Decimal(new_amount)
            revenue.status = new_status
            revenue.notes = new_notes
            revenue.updated_by = changed_by
            revenue.save()
            
            # 記錄變更日誌
            RevenueLog.objects.create(
                revenue=revenue,
                store=revenue.store,
                action='update',
                old_amount=old_amount,
                new_amount=revenue.amount,
                old_status=old_status,
                new_status=revenue.status,
                date=revenue.date,
                notes=f'金額: ${old_amount} → ${revenue.amount}, 狀態: {old_status} → {revenue.status}',
                changed_by=changed_by
            )
            
            messages.success(request, '營收資料更新成功')
            return redirect('admin_dashboard:revenue_detail', revenue_id=revenue.id)
            
        except Exception as e:
            messages.error(request, f'更新失敗: {str(e)}')
            return redirect('admin_dashboard:revenue_detail', revenue_id=revenue.id)


class RevenueCreateView(View):
    """新增營收記錄視圖"""
    
    def get(self, request):
        """顯示新增表單"""
        context = {
            'stores': Store.objects.all(),
            'status_choices': Revenue.STATUS_CHOICES,
        }
        return render(request, 'admin_dashboard/revenue_create.html', context)
    
    def post(self, request):
        """建立新的營收記錄"""
        try:
            store_id = request.POST.get('store_id')
            date_str = request.POST.get('date')
            amount = Decimal(request.POST.get('amount'))
            status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            created_by = request.POST.get('created_by', '管理員')
            
            store = get_object_or_404(Store, id=store_id)
            date = datetime.fromisoformat(date_str).date()
            
            # 檢查是否已存在相同日期的記錄
            if Revenue.objects.filter(store=store, date=date).exists():
                messages.error(request, f'{store.name} 在 {date} 的營收記錄已存在')
                return redirect('admin_dashboard:revenue_create')
            
            # 建立營收記錄
            revenue = Revenue.objects.create(
                store=store,
                date=date,
                amount=amount,
                status=status,
                notes=notes,
                created_by=created_by,
                updated_by=created_by
            )
            
            # 記錄創建日誌
            RevenueLog.objects.create(
                revenue=revenue,
                store=store,
                action='create',
                new_amount=amount,
                new_status=status,
                date=date,
                notes=f'新增營收記錄: ${amount}',
                changed_by=created_by
            )
            
            messages.success(request, '營收記錄新增成功')
            return redirect('admin_dashboard:revenue_detail', revenue_id=revenue.id)
            
        except Exception as e:
            messages.error(request, f'新增失敗: {str(e)}')
            return redirect('admin_dashboard:revenue_create')


class RevenueDeleteView(View):
    """刪除營收記錄視圖"""
    
    def post(self, request, revenue_id):
        """刪除營收記錄"""
        revenue = get_object_or_404(Revenue, id=revenue_id)
        
        try:
            deleted_by = request.POST.get('deleted_by', '管理員')
            
            # 記錄刪除日誌
            RevenueLog.objects.create(
                store=revenue.store,
                action='delete',
                old_amount=revenue.amount,
                old_status=revenue.status,
                date=revenue.date,
                notes=f'刪除營收記錄: ${revenue.amount}',
                changed_by=deleted_by
            )
            
            store_name = revenue.store.name
            date = revenue.date
            revenue.delete()
            
            messages.success(request, f'{store_name} 在 {date} 的營收記錄已刪除')
            return redirect('admin_dashboard:revenue_crud')
            
        except Exception as e:
            messages.error(request, f'刪除失敗: {str(e)}')
            return redirect('admin_dashboard:revenue_detail', revenue_id=revenue.id)


class RevenueLogView(View):
    """營收變更日誌視圖"""
    
    def get(self, request):
        """顯示營收變更日誌"""
        # 獲取篩選參數
        store_id = request.GET.get('store_id')
        action = request.GET.get('action')
        
        logs = RevenueLog.objects.select_related('store').all()
        
        if store_id:
            logs = logs.filter(store_id=store_id)
        if action:
            logs = logs.filter(action=action)
        
        logs = logs.order_by('-changed_at')[:100]  # 最近100筆記錄
        
        context = {
            'logs': logs,
            'stores': Store.objects.all(),
            'action_choices': RevenueLog.ACTION_CHOICES,
            'selected_store': store_id,
            'selected_action': action,
        }
        return render(request, 'admin_dashboard/revenue_logs.html', context)


@require_POST
def send_notification(request):
    """Send notification to stores that haven't completed orders"""
    try:
        today = timezone.now().date()

        # Find stores that haven't completed orders today
        incomplete_stores = []
        for store in Store.objects.all():
            today_order = Order.objects.filter(store=store, date=today).first()
            if not today_order or not today_order.is_completed:
                incomplete_stores.append(store)

        # In a real app, this would send Line notifications
        # For demo, we'll just return the list of stores that would be notified

        return JsonResponse({
            'success': True,
            'message': f'已發送提醒給 {len(incomplete_stores)} 家分店',
            'notified_stores': [store.name for store in incomplete_stores]
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'錯誤: {str(e)}'
        }, status=400)
